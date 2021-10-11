#!/usr/bin/python
# -*- coding: UTF-8 -*-
##################################
# AIGO Smart Scheduling - excel lib
# team CKC
##################################
import os
import openpyxl
import pandas as pd
from sys import path
path.append(os.path.dirname(__file__))
from math import ceil
from pathlib import Path
from datetime import datetime, timedelta
from common import 讀取Yaml檔
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

debug = False

def addColumns(src, generateFile, output):
    wb = openpyxl.load_workbook(src)

    ordersDF = pd.DataFrame(wb['1-8月製令單'].values, columns=['預計開工', '預計完工', '製令編號', '產品品號', '品名',
                                                           '規格', '產量', '訂單單號', '預計出貨'])
    wb2 = openpyxl.load_workbook('src/機種對照及工時表.xlsx')
    capacityDF = pd.DataFrame(wb2['產能與類別'].values)

    additionalColumns = {'製令編號': ['製令編號'], '產品品號': ['產品品號'], '品名': ['品名'], '規格': ['規格'],
                         '產量': ['產量'], '訂單單號': ['訂單單號'], '預計出貨': ['預計出貨'], '產能': ['產能'],
                         '生產時間': ['生產時間'], '類型': ['類型']}
    additionalColumns = {i: [i] for i in ordersDF.columns}
    additionalColumns['產能'] = ['產能']
    additionalColumns['生產時間'] = ['生產時間']
    additionalColumns['類型'] = ['類型']
    columns = list(ordersDF.values[0])
    columns2 = list(capacityDF.values[1])
    for row in ordersDF.values[1:]:
        quantity = int(row[columns.index('產量')])
        productNo = row[columns.index('產品品號')]
        additionalColumns['製令編號'].append(row[columns.index('製令編號')])
        additionalColumns['產品品號'].append(productNo)
        additionalColumns['品名'].append(row[columns.index('品名')])
        additionalColumns['產量'].append(quantity)
        additionalColumns['訂單單號'].append(row[columns.index('訂單單號')])
        additionalColumns['預計出貨'].append(row[columns.index('預計出貨')])
        flag = False
        if productNo == 'None':
            break
        for k, v in enumerate(capacityDF[0] == productNo):
            if v:
                flag = True
                additionalColumns['產能'].append(capacityDF.loc[k, columns2.index('工時(秒/台)')])
                additionalColumns['生產時間'].append(int(ceil(quantity/capacityDF.loc[k, columns2.index('工時(秒/台)')])))
                additionalColumns['類型'].append(capacityDF.loc[k, columns2.index('類別')])
                break
        if not flag:
            additionalColumns['產能'].append(None)
            additionalColumns['生產時間'].append(None)
            additionalColumns['類型'].append(None)
    ordersDF['產能'] = additionalColumns['產能']
    ordersDF['生產時間'] = additionalColumns['生產時間']
    ordersDF['類型'] = additionalColumns['類型']
    ordersDF = ordersDF[1:]
    if generateFile:
        ordersDF.to_excel(output / (src.stem + '完整' + src.suffix), sheet_name='1-8月製令單', columns=['預計開工', '預計完工',
                                                                                                '製令編號', '產品品號', '品名',
                                                                                                '規格', '產量', '訂單單號',
                                                                                                '預計出貨', '產能', '生產時間', '類型'], index = False)
    return ordersDF

def genWorkingHourForm(weeks, hours, dates, comments, output):
    conf = 讀取Yaml檔(Path(os.path.dirname(__file__)).parent / 'src' / 'config.yml')


    wb = Workbook()
    sheet = wb.create_sheet('待填工時表', index=0)
    row = 1

    # 調整第一列 alignments
    for col in range(11):
        sheet[f'{chr(65+col)}{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet.row_dimensions[row].height = 30
    # 待填項目: 前一筆訂單類型
    sheet[f'A{row}'] = '前一筆訂單類型'
    sheet.merge_cells(f'A{row}:E{row}')
    # 產生各工班填寫前一筆訂單類型欄位
    for shift in range(conf['製造工班數']):
        sheet.merge_cells(f'{chr(ord("F")+shift*2)}{row}:{chr(ord("F")+shift*2+1)}{row}')
    row += 1

    formColumns = ['日期', '星期', '備註', '製1~8', '新增工時']
    columns = [chr(65+i) for i in range(len(formColumns)+2*conf['製造工班數'])]
    width = [12, 6, 14, 10, 10] + [20 for i in range(2*conf['製造工班數'])]
    chinese = ['一', '二', '三', '四', '五', '六', '七', '八', '九', '十']
    sheet.row_dimensions[row].height = 30

    # 調整欄寬
    for i in range(len(columns)):
        sheet.column_dimensions[columns[i]].width = width[i]

    mergedColumns = [columns[i] for i in range(len(formColumns), len(columns), 2)]

    # 設定 column name
    for i in range(conf['製造工班數']):
        formColumns.append(f'製造{chinese[i]}班\n補正工時(分)')
        formColumns.append(f'製造{chinese[i]}班\n加班工時(小時)')
    for index, columnName in enumerate(formColumns):
        sheet[f'{columns[index]}{row}'] = columnName
        sheet[f'{columns[index]}{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    row += 1
    for index, week in enumerate(weeks):
        weekBegin = int(row)

        for index2, day in enumerate(week):

            # 填入日期
            sheet[f'A{row}'] = day.strftime('%Y/%m/%d')
            # 填入星期
            sheet[f'B{row}'] = dates[index][index2]
            # 填入備註
            sheet[f'C{row}'] = comments[index][index2]

            # 填入新增工時
            hour = int(round(hours[index][index2] * (1-conf['損耗率']), 0)) if hours[index][index2] > 0 else 0
            hour = hour-15 if dates[index][index2] == conf['清理']['每週重複於'] and hour > 0 else hour
            sheet[f'E{row}'] = hour

            # 調整該 row 對齊方式, 及未上班改為紅字
            for i in columns:
                sheet[f'{i}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                if not hour:
                    sheet[f'{i}{row}'].font = Font(color="FF0000")
            row += 1
        for i in range(5):
            sheet.merge_cells(f'D{3+i*7}:D{(i+1)*7}')
            sheet[f'D{3+i*7}'].alignment = Alignment(horizontal='center', vertical='center')
            sheet[f'D{3+i*7}'].font = Font(color='000000')
        for m in mergedColumns:
            sheet.merge_cells(f'{m}{weekBegin}:{m}{row-1}')
        row += 1
    for i in range(5):
        # 挑選製1~8生產工班, 預設為依據週次設定當週工班
        chosen = int(datetime.strptime(sheet[f'A{3+i*7}'].value,'%Y/%m/%d').strftime('%W')) % (conf['製造工班數']-1)
        sheet[f'D{3+i*7}'] = f'{chinese[chosen+1]}班'

    wb.save(output)

def readWorkingHourForm():
    conf = 讀取Yaml檔(Path(os.path.dirname(__file__)).parent / 'src' / 'config.yml')
    sheet = openpyxl.load_workbook('待填工時表單.xlsx')['待填工時表']
    info = {'週別':dict()}
    chinese = ['一', '二', '三', '四', '五', '六', '七', '八', '九', '十']

    # 讀取各製造工班補正工時
    workingLines = conf['製造工班數']
    column = 'F'
    for i in range(1, workingLines+1):
        info[f'製造{i}班'] = {'補正工時': dict(), '加班工時': dict(), '工時': dict()}
        for index, row in enumerate([3, 10, 17, 24, 31]):
            info[f'製造{i}班']['補正工時'][str(index+1)] = sheet[f'{column}{row}'].value
        column = chr(ord(column) + 2)
    # 讀取各製造工班加班工時
    column = 'G'
    for i in range(1, workingLines+1):
        for j in range(5):
            info[f'製造{i}班']['加班工時'][str(j+1)] = list()
            for k in range(6):
                if sheet[f'{column}{3+j*7+k}'].value == None:
                    info[f'製造{i}班']['加班工時'][str(j+1)].append(0)
                else:
                    info[f'製造{i}班']['加班工時'][str(j+1)].append(sheet[f'{column}{3+j*7+k}'].value)
        column = chr(ord(column) + 2)

    # 計算各工班每天總新增工時
    for i in range(5):
        for j in range(1, workingLines+1):
            info[f'製造{j}班']['工時'][str(i+1)] = dict()
            compenTime = info[f'製造{j}班']['補正工時'][str(i + 1)]
            if not compenTime:
                compenTime = 0
            for k in range(6):
                # 加班工時
                # 時轉分, 包含號費率
                overTime = int(round(info[f'製造{j}班']['加班工時'][str(i + 1)][k]*60*(1-conf['損耗率']),0))
                workingTime = sheet[f'E{3 + i * 7 + k}'].value + overTime
                if compenTime > 0 and workingTime >= compenTime:
                    workingTime -= compenTime
                    compenTime = 0
                elif compenTime > 0:
                    workingTime = 0
                    compenTime -= workingTime
                # 工時包含 每日新增工時+加班-補正工時
                info[f'製造{j}班']['工時'][str(i + 1)][sheet[f'A{3+i*7+k}'].value] = workingTime
    # 讀取前一筆訂單類型
    for shift in range(conf['製造工班數']):
        info[f'製造{shift+1}班']['前筆訂單類型'] = sheet[f'{chr(ord("F")+2*shift)}1'].value
    # 預計排程週別
    for i in range(5):
        info['週別'][str(i+1)]={'開始': datetime.strptime(sheet[f'A{3+i*7}'].value, '%Y/%m/%d')+timedelta(days=7),
                              '結束': datetime.strptime(sheet[f'A{1+7*(i+1)}'].value, '%Y/%m/%d')+timedelta(days=8),
                              '製造': chinese.index(sheet[f'D{3+7*i}'].value[0])+1}
    return info

def scheduleOutput(info):
    # ['新增工時', '每日工時', 'flag', '前筆訂單類型', '訂單', '總工時', '換線時間', '上週剩餘工時', '本週剩餘工時', '補正工時', '加班工時']x`

    conf = 讀取Yaml檔(Path(os.path.dirname(__file__)).parent / 'src' / 'config.yml')
    setupConf = 讀取Yaml檔(Path(os.path.dirname(__file__)).parent / 'src' / '換線時間.yml')
    # 創建 excel
    wb = Workbook()
    rowHeight = 18
    chinese = ['一', '二', '三', '四', '五', '六', '七', '八', '九', '十']
    # 週
    for week in range(1, 6):
        # 共同設定
        ## 開始日期, 結束日期
        startdate = list(info[str(week)][f'製造1班']['每日工時'].keys())[0].replace('/','-')
        enddate = list(info[str(week)][f'製造1班']['每日工時'].keys())[-1].replace('/','-')
        sheet = wb.create_sheet(f'{startdate}_{enddate}', index=week-1)

        week = str(week)
        # 調整欄寬
        columnsWidth = {'A':14, 'B':12, 'C':16, 'D':18, 'E':12, 'F':6, 'G':15, 'H':8, 'I':6, 'J':12, 'K':6, 'L':10, 'M':12}
        for key, value in columnsWidth.items():
            sheet.column_dimensions[key].width = value
        row = 1
        # 調整列高
        for i in range(1, conf['製造工班數']+1):
            shift = f'製造{chinese[i-1]}班'
            sheet.row_dimensions[row].height = rowHeight

            # 設定標題列
            sheet.merge_cells(f'A{row}:H{row + 1}')
            sheet[f'A{row}'] = '福佑電機/週製令明細表'
            sheet[f'A{row}'].alignment = Alignment(horizontal='center')
            sheet[f'A{row}'].font = Font(size=24)
            sheet.merge_cells(f'I{row}:M{row + 1}')
            sheet[f'I{row}'] = shift
            sheet[f'I{row}'].alignment = Alignment(horizontal='center')
            sheet[f'I{row}'].font = Font(size=24)
            row += 2
            sheet.row_dimensions[row].height = rowHeight

            # 寫入第一層 meta-data
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                sheet[f'{col}{row}'].font = Font(size=12)
                sheet[f'{col}{row}'].alignment = Alignment(vertical='center')
            sheet.merge_cells(f'A{row}:B{row}')
            sheet[f'A{row}'] = '稼動總工時'
            sheet[f'C{row}'] = info[str(week)][f'製造{i}班']['總工時']
            sheet[f'D{row}'] = '損耗率'
            sheet[f'E{row}'] = f'{conf["損耗率"] * 100}%'
            sheet.merge_cells(f'I{row}:J{row}')
            sheet[f'I{row}'] = '製令別'
            sheet.merge_cells(f'K{row}:M{row}')
            sheet[f'K{row}'] = ''
            row += 1
            # 寫入第二層 meta-data
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                sheet[f'{col}{row}'].font = Font(size=12)
                sheet[f'{col}{row}'].alignment = Alignment(vertical='center')
            sheet.merge_cells(f'A{row}:B{row}')
            sheet[f'A{row}'] = '(上)剩餘工時'
            if week != '1':
                sheet[f'C{row}'] = info[str(int(week)-1)][f'製造{i}班']['本週剩餘工時']
            else:
                sheet[f'C{row}'] = 0
            sheet[f'D{row}'] = '換線補正'
            kinds = [info[str(week)][f'製造{i}班']['前筆訂單類型']]
            setupTimes = 0
            for j in info[str(week)][f'製造{i}班']['訂單']:
                if j[-1] != kinds[-1]:
                    setupTimes += setupConf[kinds[-1]][j[-1]]
                    kinds.append(j[-1])
            sheet[f'E{row}'] = setupTimes
            sheet.merge_cells(f'F{row}:G{row}')
            sheet[f'F{row}'] = '補正工時'
            sheet[f'H{row}'] = info[str(week)][f'製造{i}班']['補正工時'] if info[str(week)][f'製造{i}班']['補正工時'] else 0
            sheet.merge_cells(f'I{row}:J{row}')
            sheet[f'I{row}'] = '製表日期'
            sheet.merge_cells(f'K{row}:M{row}')
            sheet[f'K{row}'] = datetime.now().strftime('%Y/%m/%d')
            row += 1
            # 寫入第三層 meta-data
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                sheet[f'{col}{row}'].font = Font(size=12)
                sheet[f'{col}{row}'].alignment = Alignment(vertical='center')
            sheet.row_dimensions[row].height = rowHeight
            sheet.merge_cells(f'A{row}:B{row}')
            sheet[f'A{row}'] = '(本)新增工時'
            sheet[f'C{row}'] = info[str(week)][f'製造{i}班']['新增工時']
            sheet[f'D{row}'] = '(本)剩餘工時'
            sheet[f'E{row}'] = max(0, -1*(info[str(week)][f'製造{i}班']['總工時'] - info[str(week)][f'製造{i}班']['新增工時'] - setupTimes - info[str(week)][f'製造{i}班']['補正工時']))
            sheet.merge_cells(f'F{row}:G{row}')
            sheet[f'F{row}'] = '加班工時'
            if info[str(week)][f'製造{i}班']['額外加班'] > 0:
                sheet[f'H{row}'] = f"{sum(info[str(week)][f'製造{i}班']['加班工時'])}+{info[str(week)][f'製造{i}班']['額外加班']//(60*0.97)}小時({info[str(week)][f'製造{i}班']['額外加班']}分鐘)"
            else:
                sheet[f'H{row}'] = sum(info[str(week)][f'製造{i}班']['加班工時'])
            sheet.merge_cells(f'I{row}:J{row}')
            sheet[f'I{row}'] = '備註'
            sheet.merge_cells(f'K{row}:M{row}')
            if info[str(week)][f'製造{i}班']['備註']:
                sheet[f'K{row}'] = f"{info[str(week)][f'製造{i}班']['備註']} 張製令單未排"
            else:
                sheet[f'K{row}'] = ''
            row += 1
            # 排程表欄位
            sheet.row_dimensions[row].height = rowHeight
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                sheet[f'{col}{row}'].fill = PatternFill("solid", fgColor="DDDDDD")
                sheet[f'{col}{row}'].font = Font(size=14)
                sheet[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            sheet[f'A{row}'] = '日期'
            sheet[f'B{row}'] = '製令編號'
            sheet[f'C{row}'] = '產品品號'
            sheet[f'D{row}'] = '品名'
            sheet[f'E{row}'] = '工時/分'
            sheet[f'F{row}'] = '換線'
            sheet.merge_cells(f'G{row}:H{row}')
            sheet[f'G{row}'] = '規格'
            sheet[f'I{row}'] = '數量'
            sheet[f'J{row}'] = '計畫內容'
            sheet[f'K{row}'] = '台/分'
            sheet[f'L{row}'] = '類型'
            sheet[f'M{row}'] = '預計出貨'
            row += 1

            # 為了計算各訂單之開始生產時間
            preType = info[str(week)][f'製造{i}班']['前筆訂單類型']
            lastLeft = 0 if int(week) < 2 else info[str(int(week)-1)][f'製造{i}班']['本週剩餘工時']
            sortedTimeList = sorted(info[str(week)][f'製造{i}班']['每日工時'].keys())
            num = -1
            while lastLeft > 0 and num < 5:
                num += 1
                diff = lastLeft - info[str(week)][f'製造{i}班']['每日工時'][sortedTimeList[num]]
                lastLeft = max(0, diff)
                info[str(week)][f'製造{i}班']['每日工時'][sortedTimeList[num]] = max(0, -1*diff)
            if num == -1:
                num += 1
            startDate = sortedTimeList[num]
            count = 0
            for order in info[str(week)][f'製造{i}班']['訂單']:
                sheet.row_dimensions[row].height = rowHeight
                sheet[f'A{row}'] = startDate
                count += 1
                # 計算此訂單完成時間 = 下筆訂單開工時間
                if count != len(info[str(week)][f'製造{i}班']['訂單']):
                    orderManu = int(order[-2])
                    while orderManu > 0 and num < 5:
                        diff = orderManu - info[str(week)][f'製造{i}班']['每日工時'][sortedTimeList[num]]
                        orderManu = max(0, diff)
                        info[str(week)][f'製造{i}班']['每日工時'][sortedTimeList[num]] = max(0, -1*diff)
                        num += 1
                    num -= 1
                    startDate = sortedTimeList[num]

                sheet[f'B{row}'] = order[2]
                sheet[f'C{row}'] = order[3]
                sheet[f'D{row}'] = order[4]
                sheet[f'E{row}'] = order[-2]
                sheet[f'F{row}'] = setupConf[preType][order[-1]]
                preType = order[-1]
                sheet.merge_cells(f'G{row}:H{row}')
                sheet[f'G{row}'] = order[5]
                sheet[f'I{row}'] = order[6]
                sheet[f'J{row}'] = ''
                sheet[f'K{row}'] = order[9]
                sheet[f'L{row}'] = order[11]
                sheet[f'M{row}'] = order[8]
                row += 1

            sheet.row_dimensions[row].height = 24
            sheet[f'A{row}'] = f'核准:{" "*50}製表人（生管）:{" "*40}{datetime.now().strftime("%Y/%m/%d")}'
            sheet.merge_cells(f'A{row}:M{row}')
            sheet[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
            sheet[f'A{row}'].font = Font(size=16)
            row += 3

        row += 1

    # 輸出排程表
    wb.save(f"{list(info[str(1)][f'製造1班']['每日工時'].keys())[0].replace('/','-')}-{list(info[str(5)][f'製造1班']['每日工時'].keys())[-1].replace('/','-')}排程表.xlsx")


if __name__=='__main__':
    pass
